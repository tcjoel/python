# openpyxl is a library to install to be able to read, write excel 2020 files -> pip install openpyxl
import openpyxl
inv_file = openpyxl.load_workbook("inventory.xlsx")   # load_workbook to load your file in the variable
product_list = inv_file["Sheet1"] # content the all list with row and colone for the first sheet

products_per_supplier = {}
total_value_per_supplier = {}
product_inventory_lessThan_10 = {}

for product_row in range(2, product_list.max_row + 1):  # max_row give you the number of your max row
    supplier_name = product_list.cell(product_row, 4).value  # cell fonction help to choose the content of a particular cellule
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_nub = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)
    ## calculation of number of products per supplier
    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier[supplier_name]  # current_num_products = products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] = current_num_products + 1
    else:
        # print ("adding a new supplier")
        products_per_supplier[supplier_name] = 1

    ## calculation total value of inventory per supplier
    if supplier_name in total_value_per_supplier:
        actual_total_value = total_value_per_supplier[supplier_name]
        total_value_per_supplier[supplier_name] = actual_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    ## Product number with inventory less than 10
    if int(inventory) < 10:
        product_inventory_lessThan_10[int(product_nub)] = int(inventory)

    ## adding a value to the empy 5th colone for total inventory
    inventory_price.value = inventory * price

inv_file.save("test_inventory_creation.xlsx") # to save a newly created row in a new file
print(product_inventory_lessThan_10)
print(products_per_supplier)
print(total_value_per_supplier)



