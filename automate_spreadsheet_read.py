# automate the read from a spreadsheet file to do some date processing
# in the file called inventory.xlsx we will automate some task like
# list each company with respective product count
# list product with inventory less than 10
# list each company with respective total inventory value
# calculate and write inventory value for each product into spreadsheet
# "openpyxl" is a package that help to manage spreadsheet easily "pip install openpyxl"

import openpyxl
inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

products_per_supplier = { }

# print(product_list.max_row)

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] = current_num_products + 1
    else:
        print("adding new supplier")
        products_per_supplier[supplier_name] = 1
print(products_per_supplier)
