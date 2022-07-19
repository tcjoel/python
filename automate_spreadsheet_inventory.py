# automate the read from a spreadsheet file to do some date processing
# in the file called inventory.xlsx we will automate some task like
# list each company with respective product count
# list product with inventory less than 10
# list each company with respective total inventory value
# calculate and write inventory value for each product into spreadsheet
# "openpyxl" is a package that help to manage spreadsheet easily "pip install openpyxl"

import openpyxl
inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

products_per_supplier = { }

# print(product_list.max_row)

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] = current_num_products + 1
    else:
        print("adding new supplier")
        products_per_supplier[supplier_name] = 1
print(products_per_supplier)

# now let's calculate the total inventory per company
inventory_per_supplier = { }

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    
    invantory_price = product_list.cell(product_row, 5)  #line that we want to add in a new file
    invantory_price.value = inventory*price
    
    if supplier_name in inventory_per_supplier:
        inventory_per_supplier[supplier_name] = inventory_per_supplier[supplier_name] + inventory*price
    else:
        print("adding new supplier")
        inventory_per_supplier[supplier_name] = inventory*price
print(inventory_per_supplier)

# Now let's print out all the product with inventory less than 10

product_inventory_less_10 = { }

for product_row in range(2, product_list.max_row + 1):
    product_num = product_list.cell(product_row, 1).value
    inventory = product_list.cell(product_row, 2).value
    if inventory < 10:
        product_inventory_less_10[int(product_num)] = int(inventory)
print(product_inventory_less_10)

# add value for total inventory price "inv_file.save()" help to save the file


print(products_per_supplier)
print(inventory_per_supplier)
print(product_inventory_less_10)

inv_file.save("inventory_with_total_value.xlsx")